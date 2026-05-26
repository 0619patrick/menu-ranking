"""
数据处理核心:
1. 通过 POS 适配器读取源数据 → 标准 4 列 DataFrame
2. 按指定餐厅的菜单分类填充堂食销量（排除 KT/FP）
3. 把 KT/FP 项目单独整理成外卖区（自取分上下）
4. 生成左右并列的 Excel 对照表

不再关心源数据是哪个 POS 系统、属于哪类餐厅——
全部由 (restaurant_type, pos_type) 路由到对应的 menu / adapter。
"""
import io
import re
import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.pos_adapters import get_adapter
from app.menus import get_menu
from app.menus.base import Menu


def _delivery_pattern(menu: Menu) -> str:
    """所有平台的 markers 扁平后拼成 regex（用来粗筛"是否外卖行"）"""
    return '|'.join(re.escape(m) for m in menu.all_delivery_markers)


# ============= 样式 =============
BLUE = PatternFill('solid', fgColor='FF8DB4E2')
LIGHT_BLUE = PatternFill('solid', fgColor='FFDAEEF3')
YELLOW = PatternFill('solid', fgColor='FFFFFF00')
GRAY_HEADER = PatternFill('solid', fgColor='FFA6A6A6')
LIGHT_GRAY = PatternFill('solid', fgColor='FFE7E6E6')
F2_GRAY = PatternFill('solid', fgColor='FFF2F2F2')
DELIVERY_BG = PatternFill('solid', fgColor='FFFFF2CC')

FONT_TITLE = Font(name='宋体', size=14, bold=True)
FONT_HEADER = Font(name='宋体', size=11, bold=True)
FONT_SECTION = Font(name='宋体', size=10, bold=True)
FONT_SECTION_W = Font(name='宋体', size=10, bold=True, color='FFFFFFFF')
FONT_CAT = Font(name='宋体', size=11)
FONT_DATA = Font(name='宋体', size=10)
FONT_EXTRA = Font(name='宋体', size=10, color='FF666666', italic=True)

CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')

_thin = Side(border_style='thin', color='FFBFBFBF')
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


# ============= 数据加载 =============

def load_source(file_obj, pos_type: str) -> pd.DataFrame:
    """通过指定 POS 适配器加载源数据，返回标准 4 列 DataFrame"""
    adapter = get_adapter(pos_type)
    return adapter.load(file_obj)


def apply_deletions(src: pd.DataFrame, deletions) -> pd.DataFrame:
    """
    用户在前端 ▶ 里点 × 删掉的行，从源数据里剔除后再交给后续逻辑。

    deletions: list of {'name': POS项目名, 'cat': 分类}
    """
    if not deletions:
        return src
    keys = {(d.get('name'), d.get('cat')) for d in deletions}
    if not keys:
        return src
    mask = [(n, c) not in keys
            for n, c in zip(src['项目名称'].values, src['分类'].values)]
    return src[mask]


# ============= 通用聚合（不依赖菜单） =============

def precompute_dinein_by_name(src, menu: Menu):
    """
    一次性把堂食（排除外卖分类 + 排除 drop_categories）的数据按
    (项目名称, 分类) 预聚合好，返回 dict: { POS项目名: [(分类, qty, amt), ...] }

    排除 drop_categories 是为了防止「被丢弃分类」里的项目名与正常分类重名时产生双重计算。
    例：General Meal Set Item 里的 Longan lce 与 Drinks 里的 Longan lce 同名，
    若不排除，get_dinein_sales 会把两份都加进去。

    所有 get_dinein_sales / get_dinein_sales_detail / build_dinein_extras
    都基于这个 dict 查询，把过去 O(N×K) 的全表过滤变成 O(K) 的字典查找。
    """
    pattern = _delivery_pattern(menu)
    dinein = src[~src['分类'].str.contains(pattern, na=False, regex=True)]
    # 排除 drop_categories，避免被丢弃分类的同名项目混入堂食计算
    if menu.drop_categories:
        dinein = dinein[~dinein['分类'].isin(menu.drop_categories)]
    if dinein.empty:
        return {}
    agg = dinein.groupby(['项目名称', '分类'], as_index=False).agg(
        数量=('数量', 'sum'), 金额=('金额', 'sum')
    )
    by_name = {}
    for name, cat, q, a in zip(agg['项目名称'], agg['分类'], agg['数量'], agg['金额']):
        by_name.setdefault(name, []).append((cat, int(q), int(round(a))))
    return by_name


def get_dinein_sales(name_list, by_name):
    """加总指定 POS 名列表的堂食 (数量, 金额)。O(K) 字典查找。"""
    q = a = 0
    for n in name_list:
        for _cat, qty, amt in by_name.get(n, ()):
            q += qty
            a += amt
    return q, a


def precompute_addon_split(src, menu: Menu):
    """
    四季芬芳「加料」专用：把 addon_categories 里的行按 (POS项目名) 分成
    「收费版 价>0」「内含版 价=0」两组，返回 dict:
      { POS项目名: {'paid_q', 'paid_a', 'free_q', 'free_a'} }
    没配置 addon_categories 就直接返回 {}。
    """
    if not menu.addon_categories:
        return {}
    sub = src[src['分类'].isin(menu.addon_categories)]
    if sub.empty:
        return {}
    sub = sub.copy()
    sub['is_paid'] = sub['金额'] > 0
    agg = sub.groupby(['项目名称', 'is_paid'], as_index=False).agg(
        数量=('数量', 'sum'), 金额=('金额', 'sum')
    )
    out = {}
    for name, is_paid, q, a in zip(agg['项目名称'], agg['is_paid'],
                                    agg['数量'], agg['金额']):
        entry = out.setdefault(name, {'paid_q': 0, 'paid_a': 0,
                                      'free_q': 0, 'free_a': 0})
        if is_paid:
            entry['paid_q'] = int(q)
            entry['paid_a'] = int(round(a))
        else:
            entry['free_q'] = int(q)
            entry['free_a'] = int(round(a))
    return out


def get_addon_split(name_list, addon_lookup):
    """加总指定 POS 名列表的「收费 / 内含」拆分结果。"""
    paid_q = paid_a = free_q = 0
    for n in name_list:
        d = addon_lookup.get(n)
        if not d:
            continue
        paid_q += d['paid_q']
        paid_a += d['paid_a']
        free_q += d['free_q']
    return paid_q, paid_a, free_q


def get_dinein_sales_detail(name_list, by_name):
    """
    返回 (total_qty, total_amt, variants)，按 (POS项目名, 分类) 拆。

    同名跨分类（例如「鮮椰子水」既在「飲品」也在套餐拆解分类「ODO飲品(不能撞餐)」）
    会拆成多条 variant，给前端 ▶ 展开用。
    """
    variants = []
    for n in name_list:
        for cat, qty, amt in by_name.get(n, ()):
            if qty == 0 and amt == 0:
                continue
            variants.append({'name': n, 'cat': cat, 'qty': qty, 'amt': amt})
    # 金额降序；同金额按数量降序
    variants.sort(key=lambda v: (-v['amt'], -v['qty']))
    return (
        sum(v['qty'] for v in variants),
        sum(v['amt'] for v in variants),
        variants,
    )


def _sort_extras_cats(extras):
    """
    非平台分类按金额降序; 带「平台」字样的分类作为一整块, 插到「下午茶」之后。
    没有「下午茶」分类时, 平台块放末尾。
    """
    amt = lambda c: sum(x[2] for x in extras[c])
    non_platform = sorted(
        [c for c in extras if '平台' not in c],
        key=lambda c: -amt(c),
    )
    platform = sorted(
        [c for c in extras if '平台' in c],
        key=lambda c: -amt(c),
    )
    if '下午茶' in non_platform:
        idx = non_platform.index('下午茶')
        return non_platform[: idx + 1] + platform + non_platform[idx + 1 :]
    return non_platform + platform


def merge_new_items(items_in_cat):
    """
    把 new_in_section[菜单分类] 里同名新菜（来自不同 POS 大类）合并成一行。

    入参: [(name, pos_cat, qty, amt), ...]
    返回: [(name, total_qty, total_amt, [pos_cats]), ...]，按金额降序
    """
    agg = {}  # name → [qty, amt, set(cats)]
    for name, pos_cat, q, a in items_in_cat:
        if name in agg:
            agg[name][0] += q
            agg[name][1] += a
            agg[name][2].add(pos_cat)
        else:
            agg[name] = [q, a, {pos_cat}]
    out = [(name, q, a, sorted(cats)) for name, (q, a, cats) in agg.items()]
    out.sort(key=lambda r: -r[2])
    return out


def route_unmatched_items(by_name, used_names, menu: Menu):
    """
    把所有「菜单没匹配上」的 POS 行按 menu.cat_map / force_cat / drop_names 路由：
      - 目标是某菜单分类名 → 当作🆕新菜，加进 new_in_section[菜单分类] 列表
      - 目标是 '__OUT__'   → 放进 extras（保留原 POS 大类作为段标题）
      - 目标是 '__DROP__' / drop_names 命中 / drop_categories 命中 → 丢弃

    返回 (new_in_section, extras):
      - new_in_section: { 菜单分类名: [(name, pos_cat, qty, amt), ...] }
      - extras:         { POS分类名: [(name, qty, amt), ...] }
    """
    new_in_section = {}
    extras = {}
    for name, rows in by_name.items():
        if name in used_names:
            continue
        for cat, q, a in rows:
            if cat in menu.drop_categories:
                continue
            if q == 0 and a == 0:
                continue
            target = menu.route_new_item(name, cat)
            if target == '__DROP__' or target is None:
                continue
            if target == '__OUT__':
                extras.setdefault(cat, []).append((name, q, a))
            else:
                new_in_section.setdefault(target, []).append((name, cat, q, a))
    for c in extras:
        extras[c].sort(key=lambda x: -x[2])
    return new_in_section, extras


def build_dinein_extras(by_name, used_names, menu: Menu):
    """向后兼容封装：只返回 extras 部分。新代码请用 route_unmatched_items。"""
    _, extras = route_unmatched_items(by_name, used_names, menu)
    return extras


def build_delivery(src, menu: Menu):
    """
    收集所有外卖分类的项目，按平台分组返回。

    返回结构:
      [
        {
          'platform': 'Keeta',           # menu.delivery_platforms 的 key
          'normal':   [{'cat':..., 'items':[{name,qty,amt,merged}]}, ...],
          'selftake': [...同 normal 结构, 含「自取」字样的子分类...],
          'adjustments': [               # 补差价(adjust_marker)行, 按项目名聚合
              {'name': 'Keeta 補差價', 'qty':4, 'amt':25,
               'merged': [{'name':..,'cat':..,'qty':..,'amt':..}, ...]},
              ...
          ],
        },
        ...其他平台
      ]
    平台顺序按 menu.delivery_platforms 字典声明顺序。
    """
    if src.empty:
        return []

    # 反向映射: POS 项目名 → 菜单标准菜名（用于同分类内合并显示）
    pos_to_dish = {}
    for _cat, items in menu.items:
        for dish_name, _p, _u, pos_names in items:
            for pn in pos_names:
                pos_to_dish[pn] = dish_name
    # 套餐别名（如 Set+Drink）：菜单未列，但希望用中文显示
    for pn, alias in menu.pos_aliases.items():
        pos_to_dish.setdefault(pn, alias)

    # 把整个 src 按 (平台, 是否补差价) 切片
    platforms_data = {p: {'normal_rows': [], 'selftake_rows': [], 'adjust_rows': []}
                      for p in menu.delivery_platforms.keys()}

    for cat, sub in src.groupby('分类'):
        platform = menu.classify_platform(cat)
        if platform is None:
            continue
        is_adjust = menu.is_adjust_category(cat)
        bucket = 'adjust_rows' if is_adjust else (
            'selftake_rows' if '自取' in cat else 'normal_rows'
        )
        for _, r in sub.iterrows():
            platforms_data[platform][bucket].append((cat, r['项目名称'], int(r['数量']), int(r['金额'])))

    # 渲染各平台
    result = []
    for platform in menu.delivery_platforms.keys():
        pd_data = platforms_data[platform]

        def build_section_group(rows):
            """rows: list of (cat, pos_name, qty, amt) → [{cat, items:[{name,qty,amt,merged}]}, ...]"""
            by_cat = {}
            for cat, pn, q, a in rows:
                if q == 0 and a == 0:
                    continue
                by_cat.setdefault(cat, []).append((pn, q, a))
            sections = []
            for cat in sorted(by_cat.keys(),
                              key=lambda c: -sum(a for _, _, a in by_cat[c])):
                # 同分类内按 (菜单标准名 或 POS 名) 合并
                groups = {}  # display_key → {pos_name: [q, a]}
                for pn, q, a in by_cat[cat]:
                    display_key = pos_to_dish.get(pn, pn)
                    inner = groups.setdefault(display_key, {})
                    if pn in inner:
                        inner[pn][0] += q
                        inner[pn][1] += a
                    else:
                        inner[pn] = [q, a]
                items = []
                for display_key, by_pn in groups.items():
                    variants = [{'name': pn, 'qty': q, 'amt': a}
                                for pn, (q, a) in by_pn.items()]
                    total_q = sum(v['qty'] for v in variants)
                    total_a = sum(v['amt'] for v in variants)
                    # 始终用 display_key（菜单中文显示名，未匹配的菜单项则 fallback 为 POS 名）
                    # 多变体时附 merged 明细给前端 ▶ 展开
                    display_name = display_key
                    merged = variants if len(variants) >= 2 else []
                    items.append({'name': display_name, 'qty': total_q, 'amt': total_a, 'merged': merged})
                items.sort(key=lambda r: -r['amt'])
                sections.append({'cat': cat, 'items': items})
            return sections

        normal   = build_section_group(pd_data['normal_rows'])
        selftake = build_section_group(pd_data['selftake_rows'])

        # 补差价: 按 POS 项目名聚合（每笔可能金额不同，保留 ▶ 明细）
        adj_groups = {}  # pn → list of (cat, q, a)
        for cat, pn, q, a in pd_data['adjust_rows']:
            adj_groups.setdefault(pn, []).append((cat, q, a))
        adjustments = []
        for pn, entries in adj_groups.items():
            total_q = sum(q for _, q, _ in entries)
            total_a = sum(a for _, _, a in entries)
            merged = [{'name': pn, 'cat': c, 'qty': q, 'amt': a} for c, q, a in entries]
            adjustments.append({
                'name': pn,
                'qty':  total_q,
                'amt':  total_a,
                'merged': merged if len(merged) >= 2 else [],
            })
        adjustments.sort(key=lambda r: -r['amt'])

        # 没东西的平台跳过
        if not normal and not selftake and not adjustments:
            continue
        result.append({
            'platform': platform,
            'normal':   normal,
            'selftake': selftake,
            'adjustments': adjustments,
        })
    return result


# ============= Excel 输出 =============

def build_sheet(ws, shop_name, src, menu: Menu):
    """在 sheet 里建好该店的对照表"""

    # 取应用了门店补丁后的菜单
    items_patched = menu.items_for_store(shop_name)
    used_names = menu.collect_used_names(shop_name)

    # 一次性预聚合堂食数据；后面所有 get_dinein_sales 都从这里查
    by_name = precompute_dinein_by_name(src, menu)

    # 列宽（堂食 B-G，删除了 單位 列；外賣 J-N 不变）
    widths = {
        'A': 2, 'B': 13, 'C': 4.5, 'D': 36, 'E': 8, 'F': 8, 'G': 10,
        'I': 2, 'J': 22, 'K': 4.5, 'L': 36, 'M': 8, 'N': 10
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # 行 1: 总标题
    ws.row_dimensions[1].height = 28
    ws.merge_cells('B1:N1')
    ws.cell(row=1, column=2, value=f'{menu.brand} × {shop_name} 銷量對照').font = FONT_TITLE
    ws.cell(row=1, column=2).alignment = CENTER
    for col in range(2, 15):
        ws.cell(row=1, column=col).fill = BLUE

    # 行 2: 区域标题
    ws.merge_cells('B2:G2')
    ws.cell(row=2, column=2, value='堂食 (按菜單分類)').font = FONT_HEADER
    ws.cell(row=2, column=2).alignment = CENTER
    for col in range(2, 8):
        ws.cell(row=2, column=col).fill = LIGHT_BLUE

    ws.merge_cells('J2:N2')
    ws.cell(row=2, column=10, value='外賣 (KT/FP分類)').font = FONT_SECTION_W
    ws.cell(row=2, column=10).alignment = CENTER
    for col in range(10, 15):
        ws.cell(row=2, column=col).fill = GRAY_HEADER

    # 行 3: 字段头
    dinein_headers = [('B', '分類'), ('C', '排序'), ('D', '品名'),
                      ('E', '菜單價'), ('F', '數量'), ('G', '金額')]
    for col_letter, h in dinein_headers:
        c = ws[f'{col_letter}3']
        c.value = h
        c.font = FONT_HEADER
        c.fill = BLUE
        c.alignment = CENTER
        c.border = BORDER

    delivery_headers = [('J', '外賣分類'), ('K', '排序'),
                        ('L', '品名'), ('M', '數量'), ('N', '金額')]
    for col_letter, h in delivery_headers:
        c = ws[f'{col_letter}3']
        c.value = h
        c.font = FONT_HEADER
        c.fill = GRAY_HEADER
        c.alignment = CENTER
        c.border = BORDER

    merge_ranges = []

    # ===== 写堂食区 =====
    def write_dinein_row(r, vals, fill=None, font=None):
        cols = [2, 3, 4, 5, 6, 7]   # B-G: 分類/排序/品名/菜單價/數量/金額
        for col, val in zip(cols, vals):
            c = ws.cell(row=r, column=col, value=val)
            if fill:
                c.fill = fill
            c.font = font if font else FONT_DATA
            c.alignment = LEFT if col == 4 else CENTER
            c.border = BORDER

    current_row_dinein = 4

    # 茶位（若该餐厅菜单的第一项就是「茶位」）
    if items_patched and items_patched[0][0] == '茶位':
        tea_cat, tea_items = items_patched[0]
        for idx, (name, price, unit, pos_names) in enumerate(tea_items):
            q, a = get_dinein_sales(pos_names, by_name)
            write_dinein_row(current_row_dinein,
                             [tea_cat, idx+1, name, price, q, a],
                             fill=YELLOW, font=FONT_CAT if idx == 0 else FONT_DATA)
            current_row_dinein += 1
        menu_sections = items_patched[1:]
    else:
        menu_sections = items_patched

    # 大MENU 标题
    ws.merge_cells(start_row=current_row_dinein, start_column=2,
                   end_row=current_row_dinein, end_column=7)
    ws.cell(row=current_row_dinein, column=2, value='大MENU').font = FONT_SECTION
    ws.cell(row=current_row_dinein, column=2).alignment = CENTER
    for col in range(2, 8):
        ws.cell(row=current_row_dinein, column=col).fill = LIGHT_BLUE
        ws.cell(row=current_row_dinein, column=col).border = BORDER
    current_row_dinein += 1

    # 加料专用拆分
    addon_lookup = precompute_addon_split(src, menu)

    # 一次性路由未匹配项（🆕新菜进 menu sections，菜單外的进 extras）
    new_in_section, extras = route_unmatched_items(by_name, used_names, menu)

    # 各分类(按金额降序)
    for cat, items in menu_sections:
        is_addon = menu.addon_section and cat == menu.addon_section
        items_with_sales = []
        for name, price, unit, pos_names in items:
            if is_addon:
                paid_q, paid_a, free_q = get_addon_split(pos_names, addon_lookup)
                # 显示名后附「套餐內含 N次」
                disp_name = f'{name}〔套餐內含 {free_q}次〕' if free_q > 0 else name
                items_with_sales.append((disp_name, price, unit, paid_q, paid_a))
            else:
                q, a = get_dinein_sales(pos_names, by_name)
                items_with_sales.append((name, price, unit, q, a))
        # 追加路由到本分类的项目：pos_native 段不带🆕前缀，普通段带
        if not is_addon:
            is_pos_native = cat in menu.pos_native_sections
            prefix = '' if is_pos_native else '🆕 '
            for n, q, a, _cats in merge_new_items(new_in_section.get(cat, [])):
                items_with_sales.append((f'{prefix}{n}', '', '', q, a))
        items_with_sales.sort(key=lambda x: -x[4])

        start = current_row_dinein
        for idx, (name, price, unit, q, a) in enumerate(items_with_sales):
            cat_val = cat if idx == 0 else ''
            font = FONT_CAT if idx == 0 else FONT_DATA
            write_dinein_row(current_row_dinein,
                             [cat_val, idx+1, name, price, q, a],
                             font=font)
            current_row_dinein += 1
        if len(items_with_sales) > 1:
            merge_ranges.append(f'B{start}:B{current_row_dinein-1}')
        current_row_dinein += 1  # 空一行

    # 堂食菜單外（extras 已由上面 route_unmatched_items 算好）
    if extras:
        for col in range(2, 8):
            ws.cell(row=current_row_dinein, column=col).fill = LIGHT_GRAY
            ws.cell(row=current_row_dinein, column=col).border = BORDER
        ws.merge_cells(start_row=current_row_dinein, start_column=2,
                       end_row=current_row_dinein, end_column=7)
        ws.cell(row=current_row_dinein, column=2,
                value='堂食菜單外（POS有銷售但菜單未列）').font = FONT_SECTION
        ws.cell(row=current_row_dinein, column=2).alignment = CENTER
        current_row_dinein += 1

        for src_cat in _sort_extras_cats(extras):
            items = extras[src_cat]
            start = current_row_dinein
            for idx, (name, q, a) in enumerate(items):
                cat_val = src_cat if idx == 0 else ''
                write_dinein_row(current_row_dinein,
                                 [cat_val, idx+1, name, '', q, a],
                                 fill=F2_GRAY, font=FONT_EXTRA)
                current_row_dinein += 1
            if len(items) > 1:
                merge_ranges.append(f'B{start}:B{current_row_dinein-1}')
            current_row_dinein += 1

    # ===== 写外卖区(右侧) =====
    def write_delivery_row(r, vals, font=None):
        cols = [10, 11, 12, 13, 14]
        for col, val in zip(cols, vals):
            c = ws.cell(row=r, column=col, value=val)
            c.fill = DELIVERY_BG
            c.font = font if font else FONT_DATA
            c.alignment = LEFT if col == 12 else CENTER
            c.border = BORDER

    delivery_platforms = build_delivery(src, menu)
    current_row_dlv = 4

    def write_separator(value, fill=GRAY_HEADER, font=FONT_SECTION_W):
        nonlocal current_row_dlv
        for col in range(10, 15):
            ws.cell(row=current_row_dlv, column=col).fill = fill
            ws.cell(row=current_row_dlv, column=col).border = BORDER
        ws.merge_cells(start_row=current_row_dlv, start_column=10,
                       end_row=current_row_dlv, end_column=14)
        ws.cell(row=current_row_dlv, column=10, value=value).font = font
        ws.cell(row=current_row_dlv, column=10).alignment = CENTER
        current_row_dlv += 1

    def write_sections(sections):
        nonlocal current_row_dlv
        for section in sections:
            cat = section['cat']
            items = section['items']
            start = current_row_dlv
            for idx, item in enumerate(items):
                cat_val = cat if idx == 0 else ''
                font = FONT_CAT if idx == 0 else FONT_DATA
                write_delivery_row(current_row_dlv,
                                   [cat_val, idx + 1, item['name'], item['qty'], item['amt']],
                                   font=font)
                current_row_dlv += 1
            if len(items) > 1:
                merge_ranges.append(f'J{start}:J{current_row_dlv - 1}')
            current_row_dlv += 1

    for p_idx, p in enumerate(delivery_platforms):
        # 平台标题
        write_separator(f'━━━ {p["platform"]} ━━━', fill=GRAY_HEADER, font=FONT_SECTION_W)
        # 普通段
        write_sections(p['normal'])
        # 自取
        if p['selftake']:
            write_separator('─── 自取 ───', fill=GRAY_HEADER, font=FONT_SECTION_W)
            write_sections(p['selftake'])
        # 补差价
        if p['adjustments']:
            write_separator('─── 補差價 ───', fill=GRAY_HEADER, font=FONT_SECTION_W)
            for idx, item in enumerate(p['adjustments']):
                write_delivery_row(current_row_dlv,
                                   ['', idx + 1, item['name'], item['qty'], item['amt']],
                                   font=FONT_DATA)
                current_row_dlv += 1
            current_row_dlv += 1

    # 应用合并
    for mr in merge_ranges:
        ws.merge_cells(mr)
        first_cell = mr.split(':')[0]
        ws[first_cell].alignment = CENTER


# ============= 预览数据（供前端 JS 渲染） =============

def build_preview_data(shop_name, src, menu: Menu):
    """构建预览用结构化数据（供前端渲染表格）"""
    items_patched = menu.items_for_store(shop_name)
    used_names = menu.collect_used_names(shop_name)

    # 一次性预聚合
    by_name = precompute_dinein_by_name(src, menu)

    # 茶位（若该餐厅菜单的第一项就是「茶位」）
    tea_rows = []
    if items_patched and items_patched[0][0] == '茶位':
        tea_cat, tea_items_raw = items_patched[0]
        for name, price, unit, pos_names in tea_items_raw:
            q, a, variants = get_dinein_sales_detail(pos_names, by_name)
            tea_rows.append({'name': name, 'price': price, 'unit': unit,
                             'qty': q, 'amt': a,
                             'merged': variants if len(variants) > 1 else []})
        menu_iter = items_patched[1:]
    else:
        menu_iter = items_patched

    # 加料专用拆分（仅四季芬芳目前使用 addon_categories+addon_section）
    addon_lookup = precompute_addon_split(src, menu)

    # 一次性算好未匹配项的去向：哪些菜单分类要插🆕新菜、哪些进菜單外
    new_in_section, extras = route_unmatched_items(by_name, used_names, menu)

    menu_sections = []
    for cat, items_raw in menu_iter:
        rows = []
        if menu.addon_section and cat == menu.addon_section:
            # 加料段：每项显示 收费 qty/amt + 内含 free_qty（用于品名后缀「套餐內含 N次」）
            for name, price, unit, pos_names in items_raw:
                paid_q, paid_a, free_q = get_addon_split(pos_names, addon_lookup)
                rows.append({'name': name, 'price': price, 'unit': unit,
                             'qty': paid_q, 'amt': paid_a,
                             'free_qty': free_q,
                             'merged': []})
            # 按 paid_amt 降序；同金额按 paid+free 数量降序
            rows.sort(key=lambda x: (-x['amt'], -(x['qty'] + x.get('free_qty', 0))))
        else:
            for name, price, unit, pos_names in items_raw:
                q, a, variants = get_dinein_sales_detail(pos_names, by_name)
                rows.append({'name': name, 'price': price, 'unit': unit,
                             'qty': q, 'amt': a,
                             'merged': variants if len(variants) > 1 else []})
            # 插 🆕 新菜：菜单未列、但 cat_map/force_cat 路由到本分类的项目
            # 同名跨 POS 大类合并成一行，原大类列在 pos_cat 字段（合并后是 list）
            is_pos_native = cat in menu.pos_native_sections
            for n, q, a, pos_cats in merge_new_items(new_in_section.get(cat, [])):
                row = {'name': n, 'price': '', 'unit': '',
                       'qty': q, 'amt': a,
                       'pos_cat': '/'.join(pos_cats),
                       'merged': []}
                # pos_native 段不标🆕（该段本来就以 POS 原生项为准）
                if not is_pos_native:
                    row['is_new'] = True
                rows.append(row)
            rows.sort(key=lambda x: -x['amt'])
        menu_sections.append({'cat': cat, 'items': rows,
                              'pos_native': cat in menu.pos_native_sections})

    extras_sections = [
        {'cat': c, 'items': [
            {'name': menu.pos_aliases.get(n, n), 'qty': q, 'amt': a}
            for n, q, a in extras[c]
        ]}
        for c in _sort_extras_cats(extras)
    ]

    delivery_platforms = build_delivery(src, menu)
    # 直接把 build_delivery 的输出（已按平台分组、含 normal/selftake/adjustments）
    # 透传给前端 JS，前端按平台渲染表格。

    return {
        'shop_name': shop_name,
        'brand': menu.brand,
        'tea': tea_rows,
        'menu': menu_sections,
        'extras': extras_sections,
        'delivery': delivery_platforms,
    }


# ============= 入口函数 =============

def generate_excel(parsed_shops):
    """
    旧版：单店每店一个 sheet（保留作 fallback / 测试用）。

    parsed_shops: list of (shop_name: str, src: DataFrame, menu: Menu)
    返回: BytesIO 对象
    """
    wb = Workbook()
    wb.remove(wb.active)
    for shop_name, src, menu in parsed_shops:
        ws = wb.create_sheet(shop_name)
        build_sheet(ws, shop_name, src, menu)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ============= 新版：多店并排（5列/店）按 (地区, 品牌, 月份) 分 sheet =============

def build_shop_block_data(shop_name, src, menu: Menu):
    """
    单店的 5 列结构化数据，每个 block = (category, kind, rows[])
    kind: 'tea' / 'menu' / 'extra'
    每个 row: {'name', 'qty', 'amt'}
    """
    items_patched = menu.items_for_store(shop_name)
    used_names = menu.collect_used_names(shop_name)
    by_name = precompute_dinein_by_name(src, menu)
    addon_lookup = precompute_addon_split(src, menu)

    blocks = []

    # 茶位
    if items_patched and items_patched[0][0] == '茶位':
        tea_cat, tea_items = items_patched[0]
        rows = []
        for name, _p, _u, pos_names in tea_items:
            q, a = get_dinein_sales(pos_names, by_name)
            if q == 0 and a == 0:
                continue
            rows.append({'name': name, 'qty': q, 'amt': a})
        if rows:
            blocks.append({'cat': tea_cat, 'kind': 'tea', 'rows': rows})
        menu_iter = items_patched[1:]
    else:
        menu_iter = items_patched

    # 一次性路由未匹配项：新菜进所属菜单分类(🆕)，菜單外的进 extras
    new_in_section, extras = route_unmatched_items(by_name, used_names, menu)

    # 菜单分类：保持菜单声明顺序（与官方菜单 PDF 一致），
    # 这样多店并排时同一行就是同一个分类，便于横向比较销量。
    # 分类内的菜品仍按金额降序（卖得好的排前面）。
    for cat, items_raw in menu_iter:
        is_addon = bool(menu.addon_section and cat == menu.addon_section)
        rows = []
        for name, _p, _u, pos_names in items_raw:
            if is_addon:
                paid_q, paid_a, free_q = get_addon_split(pos_names, addon_lookup)
                if paid_q == 0 and paid_a == 0 and free_q == 0:
                    continue
                disp = f'{name}〔套餐內含 {free_q}次〕' if free_q > 0 else name
                rows.append({'name': disp, 'qty': paid_q, 'amt': paid_a})
            else:
                q, a = get_dinein_sales(pos_names, by_name)
                if q == 0 and a == 0:
                    continue
                rows.append({'name': name, 'qty': q, 'amt': a})
        # 追加路由进本分类的项目：pos_native 段不打🆕，普通段打🆕
        if not is_addon:
            is_pos_native = cat in menu.pos_native_sections
            for n, q, a, _cats in merge_new_items(new_in_section.get(cat, [])):
                if is_pos_native:
                    rows.append({'name': n, 'qty': q, 'amt': a})
                else:
                    rows.append({'name': f'🆕 {n}', 'qty': q, 'amt': a, 'is_new': True})
        if not rows:
            continue
        rows.sort(key=lambda r: -r['amt'])
        blocks.append({'cat': cat, 'kind': 'menu', 'rows': rows})

    # 堂食菜單外（extras 已由上面 route_unmatched_items 算好）
    for src_cat in _sort_extras_cats(extras):
        rows = [{'name': menu.pos_aliases.get(n, n), 'qty': q, 'amt': a}
                for n, q, a in extras[src_cat]]
        if rows:
            blocks.append({'cat': src_cat, 'kind': 'extra', 'rows': rows})

    # 外賣（按平台、各分类，以 "(外賣 平台) 分类" 作类目前缀，作为菜單外延伸）
    delivery = build_delivery(src, menu)
    for p in delivery:
        platform = p['platform']
        for sec in p['normal']:
            rows = [{'name': item['name'], 'qty': item['qty'], 'amt': item['amt']}
                    for item in sec['items']]
            if rows:
                blocks.append({'cat': f'(外賣 {platform}) {sec["cat"]}',
                               'kind': 'extra', 'rows': rows})
        for sec in p['selftake']:
            rows = [{'name': item['name'], 'qty': item['qty'], 'amt': item['amt']}
                    for item in sec['items']]
            if rows:
                blocks.append({'cat': f'(外賣 {platform} 自取) {sec["cat"]}',
                               'kind': 'extra', 'rows': rows})
        if p['adjustments']:
            rows = [{'name': item['name'], 'qty': item['qty'], 'amt': item['amt']}
                    for item in p['adjustments']]
            blocks.append({'cat': f'(外賣 {platform} 補差價)',
                           'kind': 'extra', 'rows': rows})

    return blocks


def build_horizontal_sheet(ws, region, brand_short, month, year, shop_blocks_list):
    """
    把同 (region, brand, month) 的多家店并排画进一个 sheet。

    shop_blocks_list: list of (shop_name, blocks)
    布局: 每店 5 列 (分類/排序/品名/數量/金額) + 1 列间隔
    菜单分类按声明顺序 & 行级对齐（同一分类在所有店占同样的行数，不够的用空行补）。
    extras / 外賣 段每店独立写在下方（不对齐）。
    """
    n_shops = len(shop_blocks_list)
    cols_per_shop = 5
    gap = 1
    block_width = cols_per_shop + gap
    total_cols = n_shops * cols_per_shop + max(0, n_shops - 1) * gap

    # 拆每家店的 blocks: 茶位 / 菜单(按分类索引) / extras
    shops = []
    canonical_cats = []   # 菜单分类的合并顺序（首次出现顺）
    for name, blocks in shop_blocks_list:
        tea = None
        menu_by_cat = {}
        extras = []
        for b in blocks:
            if b['kind'] == 'tea':
                tea = b
            elif b['kind'] == 'menu':
                menu_by_cat[b['cat']] = b
                if b['cat'] not in canonical_cats:
                    canonical_cats.append(b['cat'])
            else:
                extras.append(b)
        shops.append({'name': name, 'tea': tea, 'menu': menu_by_cat, 'extras': extras})

    # 行 1：总标题
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1, value=f'{region}門店菜品排行{year}年{month}月')
    title_cell.font = FONT_TITLE
    title_cell.alignment = CENTER
    for col in range(1, total_cols + 1):
        ws.cell(row=1, column=col).fill = BLUE

    merge_ranges = []

    # 行 2-3：每家店的列宽 / 店名 banner / 表头
    for shop_idx, s in enumerate(shops):
        col_base = 1 + shop_idx * block_width

        widths = [12, 5, 30, 7, 9]
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(col_base + i)].width = w
        if shop_idx < n_shops - 1:
            ws.column_dimensions[get_column_letter(col_base + cols_per_shop)].width = 2

        ws.merge_cells(start_row=2, start_column=col_base,
                       end_row=2, end_column=col_base + 4)
        sn_cell = ws.cell(row=2, column=col_base, value=s['name'])
        sn_cell.font = FONT_HEADER
        sn_cell.alignment = CENTER
        for c in range(col_base, col_base + 5):
            ws.cell(row=2, column=c).fill = LIGHT_BLUE
            ws.cell(row=2, column=c).border = BORDER

        for i, h in enumerate(['分類', '排序', '品名', '數量', '金額']):
            c = ws.cell(row=3, column=col_base + i, value=h)
            c.font = FONT_HEADER
            c.fill = BLUE
            c.alignment = CENTER
            c.border = BORDER

    def write_row(r, col_base, cat_label, idx, name, qty, amt, fill, font_first, font_rest):
        cells = [(0, cat_label), (1, idx), (2, name), (3, qty), (4, amt)]
        for offset, val in cells:
            c = ws.cell(row=r, column=col_base + offset, value=val)
            if fill:
                c.fill = fill
            c.font = font_first if offset == 0 else font_rest
            c.alignment = LEFT if offset == 2 else CENTER
            c.border = BORDER

    def write_empty_cells(r, col_base, fill=None):
        """该店该分类没数据时，画 5 个空格保持表格连续。"""
        for offset in range(5):
            c = ws.cell(row=r, column=col_base + offset, value='')
            if fill:
                c.fill = fill
            c.border = BORDER

    r = 4

    # ===== 茶位段：行级对齐 =====
    max_tea = max((len(s['tea']['rows']) if s['tea'] else 0) for s in shops)
    if max_tea > 0:
        for row_idx in range(max_tea):
            for shop_idx, s in enumerate(shops):
                col_base = 1 + shop_idx * block_width
                tea = s['tea']
                if tea and row_idx < len(tea['rows']):
                    row = tea['rows'][row_idx]
                    cat_label = tea['cat'] if row_idx == 0 else ''
                    font_first = FONT_CAT if row_idx == 0 else FONT_DATA
                    write_row(r + row_idx, col_base, cat_label, row_idx + 1,
                              row['name'], row['qty'], row['amt'],
                              YELLOW, font_first, FONT_DATA)
                else:
                    write_empty_cells(r + row_idx, col_base, YELLOW)
        # 茶位分类格合并
        for shop_idx, s in enumerate(shops):
            tea = s['tea']
            if tea and len(tea['rows']) > 1:
                col_base = 1 + shop_idx * block_width
                merge_ranges.append(
                    f'{get_column_letter(col_base)}{r}:'
                    f'{get_column_letter(col_base)}{r + len(tea["rows"]) - 1}'
                )
        r += max_tea + 1   # 茶位 + 空行

    # ===== 菜单分类段：行级对齐 =====
    for cat in canonical_cats:
        max_rows = max(len(s['menu'].get(cat, {}).get('rows', [])) for s in shops)
        if max_rows == 0:
            continue
        for row_idx in range(max_rows):
            for shop_idx, s in enumerate(shops):
                col_base = 1 + shop_idx * block_width
                block = s['menu'].get(cat)
                if block and row_idx < len(block['rows']):
                    row = block['rows'][row_idx]
                    cat_label = cat if row_idx == 0 else ''
                    font_first = FONT_CAT if row_idx == 0 else FONT_DATA
                    write_row(r + row_idx, col_base, cat_label, row_idx + 1,
                              row['name'], row['qty'], row['amt'],
                              None, font_first, FONT_DATA)
                else:
                    # 该店本分类无数据：第 0 行写分类名灰显，其它空白
                    if row_idx == 0:
                        write_row(r, col_base, cat, '', '—', '', '',
                                  None, FONT_CAT, FONT_DATA)
                    else:
                        write_empty_cells(r + row_idx, col_base)
        # 分类格合并：每家店各自合并自己的实际数据行（含 0 行的也合 max_rows）
        for shop_idx, s in enumerate(shops):
            col_base = 1 + shop_idx * block_width
            block = s['menu'].get(cat)
            if block and len(block['rows']) > 1:
                merge_ranges.append(
                    f'{get_column_letter(col_base)}{r}:'
                    f'{get_column_letter(col_base)}{r + len(block["rows"]) - 1}'
                )
        r += max_rows + 1

    # ===== Extras 段：每店独立写（不对齐，因为各店 POS extras 不同）=====
    extras_start_r = r
    for shop_idx, s in enumerate(shops):
        col_base = 1 + shop_idx * block_width
        cur_r = extras_start_r
        for block in s['extras']:
            kind_start = cur_r
            for row_idx, row in enumerate(block['rows']):
                cat_label = block['cat'] if row_idx == 0 else ''
                write_row(cur_r, col_base, cat_label, row_idx + 1,
                          row['name'], row['qty'], row['amt'],
                          F2_GRAY, FONT_EXTRA, FONT_EXTRA)
                cur_r += 1
            if len(block['rows']) > 1:
                merge_ranges.append(
                    f'{get_column_letter(col_base)}{kind_start}:'
                    f'{get_column_letter(col_base)}{cur_r - 1}'
                )
            cur_r += 1

    # 应用合并
    for mr in merge_ranges:
        ws.merge_cells(mr)
        first = mr.split(':')[0]
        ws[first].alignment = CENTER

    ws.freeze_panes = 'A4'


def generate_excel_grouped(grouped):
    """
    新版多店并排 Excel 输出。

    grouped: list of {
        'region':       '香港' / '内地' / ...,
        'brand_short':  '天天',
        'month':        3,
        'year':         2026,
        'shops':        [(shop_name, src, menu), ...]
    }
    每条 group 一个 sheet，sheet 名 = f'{region}{brand_short}{month}月'
    """
    wb = Workbook()
    wb.remove(wb.active)
    seen = {}
    for g in grouped:
        sheet_name = f"{g['region']}{g['brand_short']}{g['month']}月"
        # Excel 限制 sheet 名 31 字 + 不能重复
        base = sheet_name[:28]
        if base in seen:
            seen[base] += 1
            sheet_name = f'{base}_{seen[base]}'
        else:
            seen[base] = 1
            sheet_name = base
        ws = wb.create_sheet(sheet_name)
        shop_blocks = [(name, build_shop_block_data(name, src, menu))
                       for name, src, menu in g['shops']]
        build_horizontal_sheet(ws, g['region'], g['brand_short'],
                               g['month'], g['year'], shop_blocks)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def group_shops_for_export(specs_with_data, default_year=None):
    """
    按 (region, brand_short, month) 把店铺分组。

    specs_with_data: list of dict {
        'shop_name', 'src', 'menu', 'region', 'month', 'year'(可选)
    }
    返回 list of {'region','brand_short','month','year','shops':[(name,src,menu)]}
    """
    if default_year is None:
        default_year = datetime.datetime.now().year

    buckets = {}
    order = []
    for s in specs_with_data:
        menu = s['menu']
        brand_short = menu.short_name or menu.brand
        region = (s.get('region') or '其他').strip() or '其他'
        try:
            month = int(s.get('month') or datetime.datetime.now().month)
        except (TypeError, ValueError):
            month = datetime.datetime.now().month
        year = int(s.get('year') or default_year)

        key = (region, brand_short, month, year)
        if key not in buckets:
            buckets[key] = []
            order.append(key)
        buckets[key].append((s['shop_name'], s['src'], menu))

    return [
        {
            'region':      k[0],
            'brand_short': k[1],
            'month':       k[2],
            'year':        k[3],
            'shops':       buckets[k],
        }
        for k in order
    ]


def compute_stats(src, menu: Menu, shop_name: str = None):
    """从已解析的 src + menu 算统计摘要。传 shop_name 让 STORE_OVERRIDES 生效。"""
    items_patched = menu.items_for_store(shop_name)
    used = menu.collect_used_names(shop_name)
    by_name = precompute_dinein_by_name(src, menu)

    dinein_q = dinein_a = matched_items = 0
    for cat, items in items_patched:
        for name, price, unit, pos_names in items:
            q, a = get_dinein_sales(pos_names, by_name)
            dinein_q += q
            dinein_a += a
            if q > 0 or a > 0:
                matched_items += 1

    # 新菜（路由进 menu sections 但菜单未列）+ 菜單外（route → '__OUT__'）
    new_in_section, extras = route_unmatched_items(by_name, used, menu)
    new_q = sum(x[2] for v in new_in_section.values() for x in v)
    new_a = sum(x[3] for v in new_in_section.values() for x in v)
    # 新菜计入 dinein 总数（呈现在菜单分类内，业务上属"堂食菜单内"销量）
    dinein_q += new_q
    dinein_a += new_a

    extras_q = sum(x[1] for v in extras.values() for x in v)
    extras_a = sum(x[2] for v in extras.values() for x in v)

    delivery_platforms = build_delivery(src, menu)
    # delivery_qty / delivery_amt = 所有平台的正常 + 自取（不含补差价）
    dlv_q = dlv_a = 0
    adj_q = adj_a = 0
    delivery_cats = 0
    per_platform = {}
    for p in delivery_platforms:
        p_q = p_a = 0
        for sec in p['normal'] + p['selftake']:
            for item in sec['items']:
                p_q += item['qty']
                p_a += item['amt']
            delivery_cats += 1
        for item in p['adjustments']:
            adj_q += item['qty']
            adj_a += item['amt']
        per_platform[p['platform']] = {'qty': p_q, 'amt': p_a}
        dlv_q += p_q
        dlv_a += p_a

    total_menu = sum(len(items) for cat, items in items_patched)

    return {
        'menu_total':    total_menu,
        'menu_matched':  matched_items,
        'dinein_qty':    dinein_q,
        'dinein_amt':    dinein_a,
        'new_qty':       new_q,           # 🆕 新菜数量（已含在 dinein_qty 内）
        'new_amt':       new_a,
        'extras_qty':    extras_q,
        'extras_amt':    extras_a,
        'delivery_qty':  dlv_q,
        'delivery_amt':  dlv_a,
        'delivery_cats': delivery_cats,
        'adjust_qty':    adj_q,
        'adjust_amt':    adj_a,
        'per_platform':  per_platform,
    }


# 便利封装: 单文件入口（保留给可能的脚本/测试用）
def get_stats(file_obj, restaurant_type: str, pos_type: str):
    src = load_source(file_obj, pos_type)
    menu = get_menu(restaurant_type)
    return compute_stats(src, menu)
