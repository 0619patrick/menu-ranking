"""
数据处理核心:
1. 通过 POS 适配器读取源数据 → 标准 4 列 DataFrame
2. 按指定餐厅的菜单分类填充堂食销量（排除 KT/FP）
3. 把 KT/FP 项目单独整理成外卖区（自取分上下）
4. 生成左右并列的 Excel 对照表

不再关心源数据是哪个 POS 系统、属于哪类餐厅——
全部由 (restaurant_type, pos_type) 路由到对应的 menu / adapter。
"""
import io
import re
import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.adapters import get_adapter
from app.services.menu_service import get_menu
from app.models.menu import Menu


def _delivery_pattern(menu: Menu) -> str:
    return '|'.join(re.escape(m) for m in menu.all_delivery_markers)


# ============= 样式 =============
BLUE = PatternFill('solid', fgColor='FF8DB4E2')
LIGHT_BLUE = PatternFill('solid', fgColor='FFDAEEF3')
YELLOW = PatternFill('solid', fgColor='FFFFFF00')
GRAY_HEADER = PatternFill('solid', fgColor='FFA6A6A6')
LIGHT_GRAY = PatternFill('solid', fgColor='FFE7E6E6')
F2_GRAY = PatternFill('solid', fgColor='FFF2F2F2')
DELIVERY_BG = PatternFill('solid', fgColor='FFFFF2CC')

FONT_TITLE = Font(name='宋体', size=14, bold=True)
FONT_HEADER = Font(name='宋体', size=11, bold=True)
FONT_SECTION = Font(name='宋体', size=10, bold=True)
FONT_SECTION_W = Font(name='宋体', size=10, bold=True, color='FFFFFFFF')
FONT_CAT = Font(name='宋体', size=11)
FONT_DATA = Font(name='宋体', size=10)
FONT_EXTRA = Font(name='宋体', size=10)

CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')

_thin = Side(border_style='thin', color='FFBFBFBF')
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


# ============= 数据加载 =============

def load_source(file_obj, pos_type: str) -> pd.DataFrame:
    adapter = get_adapter(pos_type)
    return adapter.load(file_obj)


def apply_deletions(src: pd.DataFrame, deletions) -> pd.DataFrame:
    if not deletions:
        return src
    keys = {(d.get('name'), d.get('cat')) for d in deletions}
    if not keys:
        return src
    mask = [(n, c) not in keys
            for n, c in zip(src['项目名称'].values, src['分类'].values)]
    return src[mask]


# ============= 通用聚合（不依赖菜单） =============

def normalize_pos_names(src, menu: Menu):
    if not menu.strip_tokens and not menu.strip_regex and not menu.pos_renames:
        return src
    s = src.copy()
    col = s['项目名称'].astype(str)
    if menu.pos_renames:
        col = col.map(lambda n: menu.pos_renames.get(n, n))
    for t in menu.strip_tokens:
        col = col.str.replace(t, '', regex=False)
    if menu.strip_regex:
        known = menu.collect_used_names() | set(col)
        pats = [re.compile(p) for p in menu.strip_regex]

        def _cond_strip(name):
            for pat in pats:
                stripped = pat.sub('', name).strip()
                if stripped != name and stripped in known:
                    return stripped
            return name

        col = col.map(_cond_strip)
    s['项目名称'] = col.str.strip()
    return s


def precompute_dinein_by_name(src, menu: Menu):
    pattern = _delivery_pattern(menu)
    dinein = src[~src['分类'].str.contains(pattern, na=False, regex=True)]
    if menu.drop_categories:
        dinein = dinein[~dinein['分类'].isin(menu.drop_categories)]
    if dinein.empty:
        return {}
    agg = dinein.groupby(['项目名称', '分类'], as_index=False).agg(
        数量=('数量', 'sum'), 金额=('金额', 'sum')
    )
    by_name = {}
    for name, cat, q, a in zip(agg['项目名称'], agg['分类'], agg['数量'], agg['金额']):
        amt = int(round(a))
        if menu.drop_zero_amount and amt == 0:
            continue
        by_name.setdefault(name, []).append((cat, int(q), amt))
    return by_name


def get_dinein_sales(name_list, by_name):
    q = a = 0
    for n in name_list:
        for _cat, qty, amt in by_name.get(n, ()):
            q += qty
            a += amt
    return q, a


def precompute_addon_split(src, menu: Menu):
    if not menu.addon_categories:
        return {}
    sub = src[src['分类'].isin(menu.addon_categories)]
    if sub.empty:
        return {}
    sub = sub.copy()
    sub['is_paid'] = sub['金额'] > 0
    agg = sub.groupby(['项目名称', 'is_paid'], as_index=False).agg(
        数量=('数量', 'sum'), 金额=('金额', 'sum')
    )
    out = {}
    for name, is_paid, q, a in zip(agg['项目名称'], agg['is_paid'],
                                    agg['数量'], agg['金额']):
        entry = out.setdefault(name, {'paid_q': 0, 'paid_a': 0,
                                      'free_q': 0, 'free_a': 0})
        if is_paid:
            entry['paid_q'] = int(q)
            entry['paid_a'] = int(round(a))
        else:
            entry['free_q'] = int(q)
            entry['free_a'] = int(round(a))
    return out


def get_addon_split(name_list, addon_lookup):
    paid_q = paid_a = free_q = 0
    for n in name_list:
        d = addon_lookup.get(n)
        if not d:
            continue
        paid_q += d['paid_q']
        paid_a += d['paid_a']
        free_q += d['free_q']
    return paid_q, paid_a, free_q


def get_dinein_sales_detail(name_list, by_name):
    variants = []
    for n in name_list:
        for cat, qty, amt in by_name.get(n, ()):
            if qty == 0 and amt == 0:
                continue
            variants.append({'name': n, 'cat': cat, 'qty': qty, 'amt': amt})
    variants.sort(key=lambda v: (-v['amt'], -v['qty']))
    return (
        sum(v['qty'] for v in variants),
        sum(v['amt'] for v in variants),
        variants,
    )


def _audit_merged(pos_names, by_name, nonzero_variants, display_name):
    trigger = len(nonzero_variants) > 1 or (
        len(nonzero_variants) == 1 and nonzero_variants[0]['name'] != display_name
    )
    return nonzero_variants if trigger else []


def _sort_extras_cats(extras):
    amt = lambda c: sum(x[2] for x in extras[c])
    non_platform = sorted(
        [c for c in extras if '平台' not in c],
        key=lambda c: -amt(c),
    )
    platform = sorted(
        [c for c in extras if '平台' in c],
        key=lambda c: -amt(c),
    )
    if '下午茶' in non_platform:
        idx = non_platform.index('下午茶')
        return non_platform[: idx + 1] + platform + non_platform[idx + 1 :]
    return non_platform + platform


def merge_new_items(items_in_cat):
    agg = {}
    for name, pos_cat, q, a in items_in_cat:
        part = {'name': name, 'cat': pos_cat, 'qty': q, 'amt': a}
        b = agg.get(name)
        if b:
            b['qty'] += q; b['amt'] += a; b['cats'].add(pos_cat); b['parts'].append(part)
        else:
            agg[name] = {'qty': q, 'amt': a, 'cats': {pos_cat}, 'parts': [part]}
    out = [(name, b['qty'], b['amt'], sorted(b['cats']), b['parts'])
           for name, b in agg.items()]
    out.sort(key=lambda r: -r[2])
    return out


_EXTRAS_SUFFIX = re.compile(r'[=$＄]?\d+$')


def _agg_by_name(items):
    bucket = {}
    for name, cat, q, a in items:
        part = {'name': name, 'cat': cat, 'qty': q, 'amt': a}
        b = bucket.get(name)
        if b:
            b['qty'] += q; b['amt'] += a; b['parts'].append(part)
        else:
            bucket[name] = {'name': name, 'qty': q, 'amt': a, 'parts': [part]}
    return list(bucket.values())


def _agg_coded(items):
    code_bucket = {}
    noncoded = []
    for name, cat, q, a in items:
        part = {'name': name, 'cat': cat, 'qty': q, 'amt': a}
        if '. ' in name:
            code, dish = name.split('. ', 1)
            b = code_bucket.get(code)
            if b:
                b['qty'] += q; b['amt'] += a; b['parts'].append(part)
                if a > b['name_amt']:
                    b['name'] = name; b['name_amt'] = a; b['dish'] = dish
            else:
                code_bucket[code] = {'name': name, 'qty': q, 'amt': a,
                                     'name_amt': a, 'parts': [part], 'dish': dish}
        else:
            noncoded.append((name, cat, q, a, part))
    dish_index = {b['dish']: b for b in code_bucket.values()}
    standalone = {}
    for name, cat, q, a, part in noncoded:
        dish = _EXTRAS_SUFFIX.sub('', name).strip()
        b = dish_index.get(dish)
        if b is not None:
            b['qty'] += q; b['amt'] += a; b['parts'].append(part)
        else:
            sb = standalone.get(name)
            if sb:
                sb['qty'] += q; sb['amt'] += a; sb['parts'].append(part)
            else:
                standalone[name] = {'name': name, 'qty': q, 'amt': a, 'parts': [part]}
    return list(code_bucket.values()) + list(standalone.values())


def pop_unmerges(by_name, unmerges):
    if not unmerges:
        return []
    keys = {(u.get('name'), u.get('cat')) for u in unmerges}
    if not keys:
        return []
    forced = []
    for name in list(by_name.keys()):
        kept = []
        for cat, q, a in by_name[name]:
            if (name, cat) in keys:
                forced.append((name, cat, q, a))
            else:
                kept.append((cat, q, a))
        if kept:
            by_name[name] = kept
        else:
            del by_name[name]
    return forced


def route_unmatched_items(by_name, used_names, menu: Menu, forced_extras=()):
    new_in_section = {}
    raw = {}
    merge_targets = set(menu.extras_merge.values())
    for name, rows in by_name.items():
        if name in used_names:
            continue
        for cat, q, a in rows:
            if cat in menu.drop_categories:
                continue
            if q == 0 and a == 0:
                continue
            target = menu.route_new_item(name, cat)
            if target == '__DROP__' or target is None:
                continue
            if target == '__OUT__':
                out_cat = menu.extras_item_merge.get(name) or menu.extras_merge.get(cat, cat)
                raw.setdefault(out_cat, []).append((name, cat, q, a))
            else:
                new_in_section.setdefault(target, []).append((name, cat, q, a))
    extras = {}
    for out_cat, items in raw.items():
        buckets = _agg_coded(items) if out_cat in merge_targets else _agg_by_name(items)
        rows = [
            (b['name'], b['qty'], b['amt'], b['parts'] if len(b['parts']) > 1 else [])
            for b in buckets
        ]
        extras[out_cat] = sorted(rows, key=lambda x: -x[2])
    if forced_extras:
        menu_cats = menu._menu_cat_names()
        touched = set()
        for name, cat, q, a in forced_extras:
            if cat in menu_cats:
                new_in_section.setdefault(cat, []).append((name, cat, q, a))
            else:
                extras.setdefault(cat, []).append((name, q, a, []))
                touched.add(cat)
        for cat in touched:
            extras[cat] = sorted(extras[cat], key=lambda x: -x[2])
    return new_in_section, extras


def build_dinein_extras(by_name, used_names, menu: Menu):
    _, extras = route_unmatched_items(by_name, used_names, menu)
    return extras


def build_delivery(src, menu: Menu):
    if src.empty:
        return []

    pos_to_dish = {}
    for _cat, items in menu.items:
        for dish_name, _p, _u, pos_names in items:
            for pn in pos_names:
                pos_to_dish[pn] = dish_name
    for pn, alias in menu.pos_aliases.items():
        pos_to_dish.setdefault(pn, alias)

    platforms_data = {p: {'normal_rows': [], 'selftake_rows': [], 'adjust_rows': []}
                      for p in menu.delivery_platforms.keys()}

    for cat, sub in src.groupby('分类'):
        platform = menu.classify_platform(cat)
        if platform is None:
            continue
        is_adjust = menu.is_adjust_category(cat)
        bucket = 'adjust_rows' if is_adjust else (
            'selftake_rows' if '自取' in cat else 'normal_rows'
        )
        for _, r in sub.iterrows():
            platforms_data[platform][bucket].append((cat, r['项目名称'], int(r['数量']), int(r['金额'])))

    result = []
    for platform in menu.delivery_platforms.keys():
        pd_data = platforms_data[platform]

        def build_section_group(rows):
            by_cat = {}
            for cat, pn, q, a in rows:
                if q == 0 and a == 0:
                    continue
                by_cat.setdefault(cat, []).append((pn, q, a))
            sections = []
            for cat in sorted(by_cat.keys(),
                              key=lambda c: -sum(a for _, _, a in by_cat[c])):
                groups = {}
                for pn, q, a in by_cat[cat]:
                    display_key = pos_to_dish.get(pn, pn)
                    inner = groups.setdefault(display_key, {})
                    if pn in inner:
                        inner[pn][0] += q
                        inner[pn][1] += a
                    else:
                        inner[pn] = [q, a]
                items = []
                for display_key, by_pn in groups.items():
                    variants = [{'name': pn, 'qty': q, 'amt': a}
                                for pn, (q, a) in by_pn.items()]
                    total_q = sum(v['qty'] for v in variants)
                    total_a = sum(v['amt'] for v in variants)
                    display_name = display_key
                    merged = variants if len(variants) >= 2 else []
                    items.append({'name': display_name, 'qty': total_q, 'amt': total_a, 'merged': merged})
                items.sort(key=lambda r: -r['amt'])
                sections.append({'cat': cat, 'items': items})
            return sections

        normal = build_section_group(pd_data['normal_rows'])
        selftake = build_section_group(pd_data['selftake_rows'])

        adj_groups = {}
        for cat, pn, q, a in pd_data['adjust_rows']:
            adj_groups.setdefault(pn, []).append((cat, q, a))
        adjustments = []
        for pn, entries in adj_groups.items():
            total_q = sum(q for _, q, _ in entries)
            total_a = sum(a for _, _, a in entries)
            merged = [{'name': pn, 'cat': c, 'qty': q, 'amt': a} for c, q, a in entries]
            adjustments.append({
                'name': pn,
                'qty': total_q,
                'amt': total_a,
                'merged': merged if len(merged) >= 2 else [],
            })
        adjustments.sort(key=lambda r: -r['amt'])

        if not normal and not selftake and not adjustments:
            continue
        result.append({
            'platform': platform,
            'normal': normal,
            'selftake': selftake,
            'adjustments': adjustments,
        })
    return result


# ============= Excel 输出 =============

def build_sheet(ws, shop_name, src, menu: Menu):
    src = normalize_pos_names(src, menu)
    items_patched = menu.items_for_store(shop_name)
    used_names = menu.collect_used_names(shop_name)
    by_name = precompute_dinein_by_name(src, menu)

    widths = {
        'A': 2, 'B': 13, 'C': 4.5, 'D': 36, 'E': 8, 'F': 8, 'G': 10,
        'I': 2, 'J': 22, 'K': 4.5, 'L': 36, 'M': 8, 'N': 10
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 28
    ws.merge_cells('B1:N1')
    ws.cell(row=1, column=2, value=f'{menu.brand} × {shop_name} 銷量對照').font = FONT_TITLE
    ws.cell(row=1, column=2).alignment = CENTER
    for col in range(2, 15):
        ws.cell(row=1, column=col).fill = BLUE

    ws.merge_cells('B2:G2')
    ws.cell(row=2, column=2, value='堂食 (按菜單分類)').font = FONT_HEADER
    ws.cell(row=2, column=2).alignment = CENTER
    for col in range(2, 8):
        ws.cell(row=2, column=col).fill = LIGHT_BLUE

    ws.merge_cells('J2:N2')
    ws.cell(row=2, column=10, value='外賣 (KT/FP分類)').font = FONT_SECTION_W
    ws.cell(row=2, column=10).alignment = CENTER
    for col in range(10, 15):
        ws.cell(row=2, column=col).fill = GRAY_HEADER

    dinein_headers = [('B', '分類'), ('C', '排序'), ('D', '品名'),
                      ('E', '菜單價'), ('F', '數量'), ('G', '金額')]
    for col_letter, h in dinein_headers:
        c = ws[f'{col_letter}3']
        c.value = h
        c.font = FONT_HEADER
        c.fill = BLUE
        c.alignment = CENTER
        c.border = BORDER

    delivery_headers = [('J', '外賣分類'), ('K', '排序'),
                        ('L', '品名'), ('M', '數量'), ('N', '金額')]
    for col_letter, h in delivery_headers:
        c = ws[f'{col_letter}3']
        c.value = h
        c.font = FONT_HEADER
        c.fill = GRAY_HEADER
        c.alignment = CENTER
        c.border = BORDER

    merge_ranges = []

    def write_dinein_row(r, vals, fill=None, font=None):
        cols = [2, 3, 4, 5, 6, 7]
        for col, val in zip(cols, vals):
            c = ws.cell(row=r, column=col, value=val)
            if fill:
                c.fill = fill
            c.font = font if font else FONT_DATA
            c.alignment = LEFT if col == 4 else CENTER
            c.border = BORDER

    current_row_dinein = 4

    if items_patched and items_patched[0][0] == '茶位':
        tea_cat, tea_items = items_patched[0]
        for idx, (name, price, unit, pos_names) in enumerate(tea_items):
            q, a = get_dinein_sales(pos_names, by_name)
            write_dinein_row(current_row_dinein,
                             [tea_cat, idx+1, name, price, q, a],
                             fill=YELLOW, font=FONT_CAT if idx == 0 else FONT_DATA)
            current_row_dinein += 1
        menu_sections = items_patched[1:]
    else:
        menu_sections = items_patched

    ws.merge_cells(start_row=current_row_dinein, start_column=2,
                   end_row=current_row_dinein, end_column=7)
    ws.cell(row=current_row_dinein, column=2, value='大MENU').font = FONT_SECTION
    ws.cell(row=current_row_dinein, column=2).alignment = CENTER
    for col in range(2, 8):
        ws.cell(row=current_row_dinein, column=col).fill = LIGHT_BLUE
        ws.cell(row=current_row_dinein, column=col).border = BORDER
    current_row_dinein += 1

    addon_lookup = precompute_addon_split(src, menu)
    new_in_section, extras = route_unmatched_items(by_name, used_names, menu)

    for cat, items in menu_sections:
        is_addon = menu.addon_section and cat == menu.addon_section
        items_with_sales = []
        for name, price, unit, pos_names in items:
            if is_addon:
                paid_q, paid_a, free_q = get_addon_split(pos_names, addon_lookup)
                disp_name = f'{name}〔套餐內含 {free_q}次〕' if free_q > 0 else name
                items_with_sales.append((disp_name, price, unit, paid_q, paid_a))
            else:
                q, a = get_dinein_sales(pos_names, by_name)
                items_with_sales.append((name, price, unit, q, a))
        if not is_addon:
            for n, q, a, _cats, _parts in merge_new_items(new_in_section.get(cat, [])):
                items_with_sales.append((n, '', '', q, a))
        items_with_sales.sort(key=lambda x: -x[4])

        start = current_row_dinein
        for idx, (name, price, unit, q, a) in enumerate(items_with_sales):
            cat_val = cat if idx == 0 else ''
            font = FONT_CAT if idx == 0 else FONT_DATA
            write_dinein_row(current_row_dinein,
                             [cat_val, idx+1, name, price, q, a],
                             font=font)
            current_row_dinein += 1
        if len(items_with_sales) > 1:
            merge_ranges.append(f'B{start}:B{current_row_dinein-1}')
        current_row_dinein += 1

    if extras:
        for col in range(2, 8):
            ws.cell(row=current_row_dinein, column=col).fill = LIGHT_GRAY
            ws.cell(row=current_row_dinein, column=col).border = BORDER
        ws.merge_cells(start_row=current_row_dinein, start_column=2,
                       end_row=current_row_dinein, end_column=7)
        ws.cell(row=current_row_dinein, column=2,
                value='堂食菜單外（POS有銷售但菜單未列）').font = FONT_SECTION
        ws.cell(row=current_row_dinein, column=2).alignment = CENTER
        current_row_dinein += 1

        for src_cat in _sort_extras_cats(extras):
            items = extras[src_cat]
            start = current_row_dinein
            for idx, (name, q, a, _m) in enumerate(items):
                cat_val = src_cat if idx == 0 else ''
                write_dinein_row(current_row_dinein,
                                 [cat_val, idx+1, name, '', q, a],
                                 fill=F2_GRAY, font=FONT_EXTRA)
                current_row_dinein += 1
            if len(items) > 1:
                merge_ranges.append(f'B{start}:B{current_row_dinein-1}')
            current_row_dinein += 1

    def write_delivery_row(r, vals, font=None):
        cols = [10, 11, 12, 13, 14]
        for col, val in zip(cols, vals):
            c = ws.cell(row=r, column=col, value=val)
            c.fill = DELIVERY_BG
            c.font = font if font else FONT_DATA
            c.alignment = LEFT if col == 12 else CENTER
            c.border = BORDER

    delivery_platforms = build_delivery(src, menu)
    current_row_dlv = 4

    def write_separator(value, fill=GRAY_HEADER, font=FONT_SECTION_W):
        nonlocal current_row_dlv
        for col in range(10, 15):
            ws.cell(row=current_row_dlv, column=col).fill = fill
            ws.cell(row=current_row_dlv, column=col).border = BORDER
        ws.merge_cells(start_row=current_row_dlv, start_column=10,
                       end_row=current_row_dlv, end_column=14)
        ws.cell(row=current_row_dlv, column=10, value=value).font = font
        ws.cell(row=current_row_dlv, column=10).alignment = CENTER
        current_row_dlv += 1

    def write_sections(sections):
        nonlocal current_row_dlv
        for section in sections:
            cat = section['cat']
            items = section['items']
            start = current_row_dlv
            for idx, item in enumerate(items):
                cat_val = cat if idx == 0 else ''
                font = FONT_CAT if idx == 0 else FONT_DATA
                write_delivery_row(current_row_dlv,
                                   [cat_val, idx + 1, item['name'], item['qty'], item['amt']],
                                   font=font)
                current_row_dlv += 1
            if len(items) > 1:
                merge_ranges.append(f'J{start}:J{current_row_dlv - 1}')
            current_row_dlv += 1

    for p_idx, p in enumerate(delivery_platforms):
        write_separator(f'━━━ {p["platform"]} ━━━', fill=GRAY_HEADER, font=FONT_SECTION_W)
        write_sections(p['normal'])
        if p['selftake']:
            write_separator('─── 自取 ───', fill=GRAY_HEADER, font=FONT_SECTION_W)
            write_sections(p['selftake'])
        if p['adjustments']:
            write_separator('─── 補差價 ───', fill=GRAY_HEADER, font=FONT_SECTION_W)
            for idx, item in enumerate(p['adjustments']):
                write_delivery_row(current_row_dlv,
                                   ['', idx + 1, item['name'], item['qty'], item['amt']],
                                   font=FONT_DATA)
                current_row_dlv += 1
            current_row_dlv += 1

    for mr in merge_ranges:
        ws.merge_cells(mr)
        first_cell = mr.split(':')[0]
        ws[first_cell].alignment = CENTER


# ============= 预览数据 =============

def build_preview_data(shop_name, src, menu: Menu):
    unmerges = src.attrs.get('unmerges') or []
    src = normalize_pos_names(src, menu)
    items_patched = menu.items_for_store(shop_name)
    used_names = menu.collect_used_names(shop_name)
    by_name = precompute_dinein_by_name(src, menu)
    forced_extras = pop_unmerges(by_name, unmerges)

    tea_rows = []
    if items_patched and items_patched[0][0] == '茶位':
        tea_cat, tea_items_raw = items_patched[0]
        for name, price, unit, pos_names in tea_items_raw:
            q, a, variants = get_dinein_sales_detail(pos_names, by_name)
            tea_rows.append({'name': name, 'price': price, 'unit': unit,
                             'qty': q, 'amt': a,
                             'merged': _audit_merged(pos_names, by_name, variants, name)})
        menu_iter = items_patched[1:]
    else:
        menu_iter = items_patched

    addon_lookup = precompute_addon_split(src, menu)
    new_in_section, extras = route_unmatched_items(by_name, used_names, menu, forced_extras)

    menu_sections = []
    for cat, items_raw in menu_iter:
        rows = []
        if menu.addon_section and cat == menu.addon_section:
            for name, price, unit, pos_names in items_raw:
                paid_q, paid_a, free_q = get_addon_split(pos_names, addon_lookup)
                rows.append({'name': name, 'price': price, 'unit': unit,
                             'qty': paid_q, 'amt': paid_a,
                             'free_qty': free_q, 'merged': []})
            rows.sort(key=lambda x: (-x['amt'], -(x['qty'] + x.get('free_qty', 0))))
        else:
            for name, price, unit, pos_names in items_raw:
                q, a, variants = get_dinein_sales_detail(pos_names, by_name)
                rows.append({'name': name, 'price': price, 'unit': unit,
                             'qty': q, 'amt': a,
                             'merged': _audit_merged(pos_names, by_name, variants, name)})
            is_pos_native = cat in menu.pos_native_sections
            for n, q, a, pos_cats, parts in merge_new_items(new_in_section.get(cat, [])):
                row = {'name': n, 'price': '', 'unit': '',
                       'qty': q, 'amt': a,
                       'pos_cat': '/'.join(pos_cats),
                       'merged': parts if len(parts) > 1 else []}
                if not is_pos_native:
                    row['is_new'] = True
                rows.append(row)
            rows.sort(key=lambda x: -x['amt'])
        menu_sections.append({'cat': cat, 'items': rows,
                              'pos_native': cat in menu.pos_native_sections})

    extras_sections = [
        {'cat': c, 'items': [
            {'name': menu.pos_aliases.get(n, n), 'qty': q, 'amt': a, 'merged': m}
            for n, q, a, m in extras[c]
        ]}
        for c in _sort_extras_cats(extras)
    ]

    delivery_platforms = build_delivery(src, menu)

    return {
        'shop_name': shop_name,
        'brand': menu.brand,
        'tea': tea_rows,
        'menu': menu_sections,
        'extras': extras_sections,
        'delivery': delivery_platforms,
    }


# ============= Excel 入口 =============

def generate_excel(parsed_shops):
    wb = Workbook()
    wb.remove(wb.active)
    for shop_name, src, menu in parsed_shops:
        ws = wb.create_sheet(shop_name)
        build_sheet(ws, shop_name, src, menu)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ============= 新版：多店并排 =============

def build_shop_block_data(shop_name, src, menu: Menu):
    unmerges = src.attrs.get('unmerges') or []
    src = normalize_pos_names(src, menu)
    items_patched = menu.items_for_store(shop_name)
    used_names = menu.collect_used_names(shop_name)
    by_name = precompute_dinein_by_name(src, menu)
    forced_extras = pop_unmerges(by_name, unmerges)
    addon_lookup = precompute_addon_split(src, menu)

    blocks = []

    if items_patched and items_patched[0][0] == '茶位':
        tea_cat, tea_items = items_patched[0]
        rows = []
        for name, _p, _u, pos_names in tea_items:
            q, a = get_dinein_sales(pos_names, by_name)
            if q == 0 and a == 0:
                continue
            rows.append({'name': name, 'qty': q, 'amt': a})
        if rows:
            blocks.append({'cat': tea_cat, 'kind': 'tea', 'rows': rows})
        menu_iter = items_patched[1:]
    else:
        menu_iter = items_patched

    new_in_section, extras = route_unmatched_items(by_name, used_names, menu, forced_extras)

    for cat, items_raw in menu_iter:
        is_addon = bool(menu.addon_section and cat == menu.addon_section)
        rows = []
        for name, _p, _u, pos_names in items_raw:
            if is_addon:
                paid_q, paid_a, free_q = get_addon_split(pos_names, addon_lookup)
                if paid_q == 0 and paid_a == 0 and free_q == 0:
                    continue
                disp = f'{name}〔套餐內含 {free_q}次〕' if free_q > 0 else name
                rows.append({'name': disp, 'qty': paid_q, 'amt': paid_a})
            else:
                q, a = get_dinein_sales(pos_names, by_name)
                if q == 0 and a == 0:
                    continue
                rows.append({'name': name, 'qty': q, 'amt': a})
        if not is_addon:
            is_pos_native = cat in menu.pos_native_sections
            for n, q, a, _cats, _parts in merge_new_items(new_in_section.get(cat, [])):
                if is_pos_native:
                    rows.append({'name': n, 'qty': q, 'amt': a})
                else:
                    rows.append({'name': n, 'qty': q, 'amt': a, 'is_new': True})
        if not rows:
            continue
        rows.sort(key=lambda r: -r['amt'])
        blocks.append({'cat': cat, 'kind': 'menu', 'rows': rows})

    for src_cat in _sort_extras_cats(extras):
        rows = [{'name': menu.pos_aliases.get(n, n), 'qty': q, 'amt': a}
                for n, q, a, _m in extras[src_cat]]
        if rows:
            blocks.append({'cat': src_cat, 'kind': 'extra', 'rows': rows})

    delivery = build_delivery(src, menu)
    for p in delivery:
        platform = p['platform']
        for sec in p['normal']:
            rows = [{'name': item['name'], 'qty': item['qty'], 'amt': item['amt']}
                    for item in sec['items']]
            if rows:
                blocks.append({'cat': f'(外賣 {platform}) {sec["cat"]}',
                               'kind': 'extra', 'rows': rows})
        for sec in p['selftake']:
            rows = [{'name': item['name'], 'qty': item['qty'], 'amt': item['amt']}
                    for item in sec['items']]
            if rows:
                blocks.append({'cat': f'(外賣 {platform} 自取) {sec["cat"]}',
                               'kind': 'extra', 'rows': rows})
        if p['adjustments']:
            rows = [{'name': item['name'], 'qty': item['qty'], 'amt': item['amt']}
                    for item in p['adjustments']]
            blocks.append({'cat': f'(外賣 {platform} 補差價)',
                           'kind': 'extra', 'rows': rows})

    return blocks


def build_horizontal_sheet(ws, region, brand_short, month, year, shop_blocks_list):
    n_shops = len(shop_blocks_list)
    cols_per_shop = 5
    gap = 1
    block_width = cols_per_shop + gap
    total_cols = n_shops * cols_per_shop + max(0, n_shops - 1) * gap

    shops = []
    canonical_cats = []
    for name, blocks in shop_blocks_list:
        tea = None
        menu_by_cat = {}
        extras = []
        for b in blocks:
            if b['kind'] == 'tea':
                tea = b
            elif b['kind'] == 'menu':
                menu_by_cat[b['cat']] = b
                if b['cat'] not in canonical_cats:
                    canonical_cats.append(b['cat'])
            else:
                extras.append(b)
        shops.append({'name': name, 'tea': tea, 'menu': menu_by_cat, 'extras': extras})

    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1, value=f'{region}門店菜品排行{year}年{month}月')
    title_cell.font = FONT_TITLE
    title_cell.alignment = CENTER
    for col in range(1, total_cols + 1):
        ws.cell(row=1, column=col).fill = BLUE

    merge_ranges = []

    for shop_idx, s in enumerate(shops):
        col_base = 1 + shop_idx * block_width
        widths = [12, 5, 30, 7, 9]
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(col_base + i)].width = w
        if shop_idx < n_shops - 1:
            ws.column_dimensions[get_column_letter(col_base + cols_per_shop)].width = 2

        ws.merge_cells(start_row=2, start_column=col_base,
                       end_row=2, end_column=col_base + 4)
        sn_cell = ws.cell(row=2, column=col_base, value=s['name'])
        sn_cell.font = FONT_HEADER
        sn_cell.alignment = CENTER
        for c in range(col_base, col_base + 5):
            ws.cell(row=2, column=c).fill = LIGHT_BLUE
            ws.cell(row=2, column=c).border = BORDER

        for i, h in enumerate(['分類', '排序', '品名', '數量', '金額']):
            c = ws.cell(row=3, column=col_base + i, value=h)
            c.font = FONT_HEADER
            c.fill = BLUE
            c.alignment = CENTER
            c.border = BORDER

    def write_row(r, col_base, cat_label, idx, name, qty, amt, fill, font_first, font_rest):
        cells = [(0, cat_label), (1, idx), (2, name), (3, qty), (4, amt)]
        for offset, val in cells:
            c = ws.cell(row=r, column=col_base + offset, value=val)
            if fill:
                c.fill = fill
            c.font = font_first if offset == 0 else font_rest
            c.alignment = LEFT if offset == 2 else CENTER
            c.border = BORDER

    def write_empty_cells(r, col_base, fill=None):
        for offset in range(5):
            c = ws.cell(row=r, column=col_base + offset, value='')
            if fill:
                c.fill = fill
            c.border = BORDER

    r = 4

    max_tea = max((len(s['tea']['rows']) if s['tea'] else 0) for s in shops)
    if max_tea > 0:
        for row_idx in range(max_tea):
            for shop_idx, s in enumerate(shops):
                col_base = 1 + shop_idx * block_width
                tea = s['tea']
                if tea and row_idx < len(tea['rows']):
                    row = tea['rows'][row_idx]
                    cat_label = tea['cat'] if row_idx == 0 else ''
                    font_first = FONT_CAT if row_idx == 0 else FONT_DATA
                    write_row(r + row_idx, col_base, cat_label, row_idx + 1,
                              row['name'], row['qty'], row['amt'],
                              YELLOW, font_first, FONT_DATA)
                else:
                    write_empty_cells(r + row_idx, col_base, YELLOW)
        for shop_idx, s in enumerate(shops):
            tea = s['tea']
            if tea and len(tea['rows']) > 1:
                col_base = 1 + shop_idx * block_width
                merge_ranges.append(
                    f'{get_column_letter(col_base)}{r}:'
                    f'{get_column_letter(col_base)}{r + len(tea["rows"]) - 1}'
                )
        r += max_tea + 1

    for cat in canonical_cats:
        max_rows = max(len(s['menu'].get(cat, {}).get('rows', [])) for s in shops)
        if max_rows == 0:
            continue
        for row_idx in range(max_rows):
            for shop_idx, s in enumerate(shops):
                col_base = 1 + shop_idx * block_width
                block = s['menu'].get(cat)
                if block and row_idx < len(block['rows']):
                    row = block['rows'][row_idx]
                    cat_label = cat if row_idx == 0 else ''
                    font_first = FONT_CAT if row_idx == 0 else FONT_DATA
                    write_row(r + row_idx, col_base, cat_label, row_idx + 1,
                              row['name'], row['qty'], row['amt'],
                              None, font_first, FONT_DATA)
                else:
                    if row_idx == 0:
                        write_row(r, col_base, cat, '', '—', '', '',
                                  None, FONT_CAT, FONT_DATA)
                    else:
                        write_empty_cells(r + row_idx, col_base)
        for shop_idx, s in enumerate(shops):
            col_base = 1 + shop_idx * block_width
            block = s['menu'].get(cat)
            if block and len(block['rows']) > 1:
                merge_ranges.append(
                    f'{get_column_letter(col_base)}{r}:'
                    f'{get_column_letter(col_base)}{r + len(block["rows"]) - 1}'
                )
        r += max_rows + 1

    extras_start_r = r
    for shop_idx, s in enumerate(shops):
        col_base = 1 + shop_idx * block_width
        cur_r = extras_start_r
        for block in s['extras']:
            kind_start = cur_r
            for row_idx, row in enumerate(block['rows']):
                cat_label = block['cat'] if row_idx == 0 else ''
                write_row(cur_r, col_base, cat_label, row_idx + 1,
                          row['name'], row['qty'], row['amt'],
                          F2_GRAY, FONT_EXTRA, FONT_EXTRA)
                cur_r += 1
            if len(block['rows']) > 1:
                merge_ranges.append(
                    f'{get_column_letter(col_base)}{kind_start}:'
                    f'{get_column_letter(col_base)}{cur_r - 1}'
                )
            cur_r += 1

    for mr in merge_ranges:
        ws.merge_cells(mr)
        first = mr.split(':')[0]
        ws[first].alignment = CENTER

    ws.freeze_panes = 'A4'


def generate_excel_grouped(grouped):
    wb = Workbook()
    wb.remove(wb.active)
    seen = {}
    for g in grouped:
        sheet_name = f"{g['region']}{g['brand_short']}{g['month']}月"
        base = sheet_name[:28]
        if base in seen:
            seen[base] += 1
            sheet_name = f'{base}_{seen[base]}'
        else:
            seen[base] = 1
            sheet_name = base
        ws = wb.create_sheet(sheet_name)
        shop_blocks = [(name, build_shop_block_data(name, src, menu))
                       for name, src, menu in g['shops']]
        build_horizontal_sheet(ws, g['region'], g['brand_short'],
                               g['month'], g['year'], shop_blocks)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_excel_by_brand(grouped):
    import zipfile
    today = datetime.datetime.now().strftime('%Y%m%d')

    by_brand = {}
    brand_order = []
    for g in grouped:
        brand = g['brand_short']
        if brand not in by_brand:
            by_brand[brand] = []
            brand_order.append(brand)
        by_brand[brand].append(g)

    if len(brand_order) == 1:
        brand = brand_order[0]
        excel_io = generate_excel_grouped(by_brand[brand])
        filename = f'{brand}_銷量對照表_{today}.xlsx'
        mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return excel_io, filename, mime

    zip_io = io.BytesIO()
    with zipfile.ZipFile(zip_io, 'w', zipfile.ZIP_DEFLATED) as zf:
        for brand in brand_order:
            excel_io = generate_excel_grouped(by_brand[brand])
            zf.writestr(f'{brand}_銷量對照表_{today}.xlsx', excel_io.getvalue())
    zip_io.seek(0)
    return zip_io, f'銷量對照表_{today}.zip', 'application/zip'


def group_shops_for_export(specs_with_data, default_year=None):
    if default_year is None:
        default_year = datetime.datetime.now().year

    buckets = {}
    order = []
    for s in specs_with_data:
        menu = s['menu']
        brand_short = menu.short_name or menu.brand
        region = (s.get('region') or '其他').strip() or '其他'
        try:
            month = int(s.get('month') or datetime.datetime.now().month)
        except (TypeError, ValueError):
            month = datetime.datetime.now().month
        year = int(s.get('year') or default_year)

        key = (region, brand_short, month, year)
        if key not in buckets:
            buckets[key] = []
            order.append(key)
        buckets[key].append((s['shop_name'], s['src'], menu))

    REGION_ORDER = {'内地': 0, '香港': 1}
    brand_first_seen = {}
    for k in order:
        b = k[1]
        if b not in brand_first_seen:
            brand_first_seen[b] = len(brand_first_seen)

    def sort_key(k):
        region, brand, month, year = k
        return (brand_first_seen[brand], REGION_ORDER.get(region, 9), year, month)

    sorted_keys = sorted(order, key=sort_key)
    return [
        {
            'region':      k[0],
            'brand_short': k[1],
            'month':       k[2],
            'year':        k[3],
            'shops':       buckets[k],
        }
        for k in sorted_keys
    ]


# ============= 统计 =============

def compute_stats(src, menu: Menu, shop_name: str = None):
    items_patched = menu.items_for_store(shop_name)
    used = menu.collect_used_names(shop_name)
    by_name = precompute_dinein_by_name(src, menu)

    dinein_q = dinein_a = matched_items = 0
    for cat, items in items_patched:
        for name, price, unit, pos_names in items:
            q, a = get_dinein_sales(pos_names, by_name)
            dinein_q += q
            dinein_a += a
            if q > 0 or a > 0:
                matched_items += 1

    new_in_section, extras = route_unmatched_items(by_name, used, menu)
    new_q = sum(x[2] for v in new_in_section.values() for x in v)
    new_a = sum(x[3] for v in new_in_section.values() for x in v)
    dinein_q += new_q
    dinein_a += new_a

    extras_q = sum(x[1] for v in extras.values() for x in v)
    extras_a = sum(x[2] for v in extras.values() for x in v)

    delivery_platforms = build_delivery(src, menu)
    dlv_q = dlv_a = 0
    adj_q = adj_a = 0
    delivery_cats = 0
    per_platform = {}
    for p in delivery_platforms:
        p_q = p_a = 0
        for sec in p['normal'] + p['selftake']:
            for item in sec['items']:
                p_q += item['qty']
                p_a += item['amt']
            delivery_cats += 1
        for item in p['adjustments']:
            adj_q += item['qty']
            adj_a += item['amt']
        per_platform[p['platform']] = {'qty': p_q, 'amt': p_a}
        dlv_q += p_q
        dlv_a += p_a

    total_menu = sum(len(items) for cat, items in items_patched)

    return {
        'menu_total':    total_menu,
        'menu_matched':  matched_items,
        'dinein_qty':    dinein_q,
        'dinein_amt':    dinein_a,
        'new_qty':       new_q,
        'new_amt':       new_a,
        'extras_qty':    extras_q,
        'extras_amt':    extras_a,
        'delivery_qty':  dlv_q,
        'delivery_amt':  dlv_a,
        'delivery_cats': delivery_cats,
        'adjust_qty':    adj_q,
        'adjust_amt':    adj_a,
        'per_platform':  per_platform,
    }


def get_stats(file_obj, restaurant_type: str, pos_type: str):
    src = load_source(file_obj, pos_type)
    menu = get_menu(restaurant_type)
    return compute_stats(src, menu)
